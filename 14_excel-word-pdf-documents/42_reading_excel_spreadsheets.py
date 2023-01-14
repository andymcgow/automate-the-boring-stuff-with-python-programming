import openpyxl
import os

# Load a workbook: -
workbook = openpyxl.load_workbook("example.xlsx")
print(type(workbook))

# Retrive the names of all the sheets in workbook: -
print(workbook.get_sheet_names())

# Store a sheet to a variable: -
sheet = workbook.get_sheet_by_name("Sheet1")

# To retrive a cell: -
cell = sheet["A1"]
# This evaluates to a cell object which has a member variable called value which contains the actual value from that cell: -
print(cell.value)
# As this is a date, it's stored in a date format, to return this as a string: -
print(str(cell.value))
# Or alternatively, instead of saving it into a variable, call: -
print(str(sheet["A1"].value))
# This cell contains a string value so you can just return the raw value: -
print(sheet["B1"].value)
# This cell contains number date so it will return an integer: -
print(sheet["C1"].value)
# And again this can be returned as string by passing str() : -
print(str(sheet["C1"].value))
# If you need to specify row and cell seperately: -
print(sheet.cell(row=1, column=2))
# This might help if you need to code something like this: -
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
