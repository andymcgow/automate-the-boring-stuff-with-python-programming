import openpyxl

# Instead of opening a workbook, openpyxl can create new Excel files: -
wb = openpyxl.Workbook()
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name("Sheet")

# Writing data to the Excel file: -
sheet["A1"] = 42
sheet["A2"] = "Hello"

# At the moment the Excel spreadsheet only exists in the computers memory and isn't written to disk.
# To write the file to disk, call wb.save() : -
wb.save("test.xlsx")
# NOTE If you ever open and existing spreadsheet and make changes/amends to that file, it's good practice to SAVE AS a new file.
# By retaining the original file, if your program creates an error and corrupts the file you will still have a backup of the original data!

# To add new sheet to your workbook called .create_sheet() : -
sheet_2 = wb.create_sheet()

# To change the name of a sheet: -
sheet_2.title = "My New Sheet Name"

wb.save("test.xlsx")

# To create a new sheet as the first sheet: -
wb.create_sheet(index=0, title="My Other Sheet")
wb.save("test.xlsx")
